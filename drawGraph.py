from Setting import*
import xlwings as xw
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)

Basic = Setting()

class drawGraph() :
    def __init__(self) : 
        pass

# x축, y축, 시트이름데이터를 행열좌표로 대입할 경우

    def property(self) :
        global xRow, xCol, yRow, yCol, sheetRow, sheetCol, maxRow, maxCol, acitveFile, activeSheet
        acitveFile, activeSheet = Basic.openFile()
        print('매트릭스값 입력')
        xRow, xCol = Basic.settingX_coordi()
        yRow, yCol = Basic.settingY_coordi()
        sheetRow, sheetCol = Basic.sheetName_coordi()
        maxRow, maxCol = Basic.Max()


    def addNewSheet(self,sheet_row, sheet_col) :
        global sheetNameCell, new_sheet
        sheetNameCell = activeSheet.cell(sheet_row,sheet_col).value
        # if sheetNameCell == None : 
        new_sheet = acitveFile.create_chartsheet(title = str(sheetNameCell))
        return new_sheet


    def addNewgraph(self) :
        for chart_num in range(xCol, maxCol + 1) :
            chart = ScatterChart()

            self.addNewSheet(sheetRow, chart_num + sheetCol - xCol)
            chart.title = f"{sheetNameCell}"
            chart.style = 13
            chart.x_axis.title = 'displacement'
            chart.y_axis.title = 'load'

            yValues = Reference(activeSheet, min_col = yCol, min_row = yRow, max_row = maxRow)
            xValues = Reference(activeSheet, min_col = chart_num, min_row = xRow, max_row = maxRow)
            series = Series(yValues, xValues, title_from_data = True)
            chart.series.append(series)
            # chart.width = 35
            # chart.height = 25
            new_sheet.add_chart(chart)

        
        saveFilename = input("저장할 파일이름 입력 : ")
        acitveFile.save(saveFilename)
            




# wb = load_workbook(filename = 'sihyon0829_2.xlsx')
# ws = wb.active


# for chart_num in range(1,5) :
#     chart = ScatterChart()
#     # chart.title = f"Scatter Chart {Reference(ws, min_col = chart_num + 3, min_row = 2)}"
#     cell_value = ws.cell(row=2, column=chart_num + 3).value
#     chart.title = f"Scatter Chart {cell_value}"
#     chart.style = 13
#     chart.x_axis.title = 'displacement'
#     chart.y_axis.title = 'load'

#     yValues = Reference(ws, min_col = 3, min_row = 2, max_row = 53)
#     xValues = Reference(ws, min_col= chart_num + 3, min_row = 3, max_row = 53)
#     series = Series(yValues, xValues, title_from_data=True)
#     chart.series.append(series)

#     ws.add_chart(chart, f"A{10 + (chart_num - 1) * 20}")

# wb.save("scatter.xlsx")
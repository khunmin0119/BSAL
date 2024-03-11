import xlwings as xw
from openpyxl import load_workbook
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)

class Setting() :

    def __init__(self):
        pass


    def openFile(self) :
        global fileForGraph,fileForGraph_sheet
        userFile = input("파일 이름 입력")
        fileForGraph = load_workbook(filename = userFile)
        fileForGraph_sheet = fileForGraph.worksheets[0]
        return fileForGraph, fileForGraph_sheet
    

    def settingX_coordi(self) :
        x_row = int(input('원하는 x축의 첫 "행 번호" 입력 : '))
        x_col = int(input('원하는 x축의 첫 "열 번호" 입력 : '))
        return x_row, x_col


    def settingY_coordi(self) :
        y_row = int(input('원하는 y축의 첫 "행 번호" 입력 : ')) - 1 
        y_col = int(input('원하는 y축의 첫 "열 번호" 입력 : '))
        return y_row, y_col


    def sheetName_coordi(self) :
        sheet_row = int(input('항목이름 시작 행 번호'))
        sheet_col = int(input('항목이름 시작 열 번호'))
        return sheet_row, sheet_col


    def Max(self) : # 행과 열의 최댓값 확인
        max_row = fileForGraph_sheet.max_row
        max_col = fileForGraph_sheet.max_column
        return max_row, max_col
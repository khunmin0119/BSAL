from Setting import*
import xlwings as xw
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)
# 
Basic = Setting()

class drawGraph2() :
    def __init__(self) : 
        pass

# 위 두줄, 왼쪽 두줄을 비웠을 경우로 작성
# load데이터 옆에 바로 변형률 데이터가 와야함(C3에 load데이터, D3에 strain데이터)
    
    def property(self) :
        global xRow, xCol, yRow, yCol, sheetRow, sheetCol, maxRow, maxCol, acitveFile, activeSheet
        acitveFile, activeSheet = Basic.openFile()
        xRow, xCol = 3, 4 
        yRow, yCol = 3, 3
        sheetRow, sheetCol = 2, 4
        maxRow, maxCol = activeSheet.max_row, activeSheet.max_column


    def addNewSheet(self,sheet_row, sheet_col) :
        global sheetNameCell, new_sheet
        sheetNameCell = activeSheet.cell(sheet_row,sheet_col).value
        new_sheet = acitveFile.create_chartsheet(title = sheetNameCell)
        return new_sheet


    def addNewgraph(self) :
        for chart_num in range(xCol, maxCol + 1) :
            chart = ScatterChart()

            self.addNewSheet(sheetRow, chart_num + sheetCol - xCol)
            chart.title = f"{sheetNameCell}"
            chart.style = 13
            chart.x_axis.title = 'displacement'
            chart.y_axis.title = 'load'

            yValues = Reference(activeSheet, min_col = yCol, min_row = yRow - 1, max_row = maxRow)
            xValues = Reference(activeSheet, min_col = chart_num, min_row = xRow, max_row = maxRow)
            series = Series(yValues, xValues, title_from_data = True)
            chart.series.append(series)
            new_sheet.add_chart(chart)
        
        saveFilename = input("저장할 파일이름 입력 : ")
        acitveFile.save(saveFilename)
            




# wb = load_workbook(filename = 'sihyon0829_2.xlsx')
# ws = wb.active


# for chart_num in range(1,5) :
#     chart = ScatterChart()
#     # chart.title = f"Scatter Chart {Reference(ws, min_col = chart_num + 3, min_row = 2)}"
#     cell_value = ws.cell(row=2, column=chart_num + 3).value
#     chart.title = f"Scatter Chart {cell_value}"
#     chart.style = 13
#     chart.x_axis.title = 'displacement'
#     chart.y_axis.title = 'load'

#     yValues = Reference(ws, min_col = 3, min_row = 2, max_row = 53)
#     xValues = Reference(ws, min_col= chart_num + 3, min_row = 3, max_row = 53)
#     series = Series(yValues, xValues, title_from_data=True)
#     chart.series.append(series)

#     ws.add_chart(chart, f"A{10 + (chart_num - 1) * 20}")

# wb.save("scatter.xlsx")